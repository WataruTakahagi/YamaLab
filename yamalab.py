import os
import openpyxl
from openpyxl import Workbook
import time
import pandas as pd
from pymeasure.instruments.keithley import Keithley2400
from datetime import datetime
import sys

def duplicate_rename(file_path):
    if os.path.exists(file_path):
        name, ext = os.path.splitext(file_path)
        i = 1
        while True:
            new_name = "{} ({:0=2}){}".format(name, i, ext)
            if not os.path.exists(new_name):
                return new_name
            i += 1
    else:
        return file_path

class cyclic_voltammetry():
    def __init__(self, setting):
        self.date = int(datetime.now().strftime("%Y%m%d"))-20000000
        #self.keithley = Keithley2400("GPIB::24")
        self.book = openpyxl.Workbook()
        self.sheet = self.book.worksheets[0]
        self.x = [] # voltage
        self.y = [] # current
        self.z = [] # time
        self.setting = pd.read_csv(setting)
        self.sample_name = self.setting['Value'][0] # ここに入力したサンプル名がファイル名に反映される。
        self.high_V = float(self.setting['Value'][1]) # high voltage (V)
        self.low_V = float(self.setting['Value'][2]) # low voltage (V)
        self.start_V = float(self.setting['Value'][3]) # 必ずしも start_V = low_V ではないが、今回は簡単のため一致させている。
        self.finish_V = float(self.setting['Value'][4])
        self.direction = 'n' # p/nの設定で掃引方向を設定できるようにしましょう。
        self.scan_rate = float(self.setting['Value'][5]) # scan rate(V/s)
        self.segments = int(self.setting['Value'][6]) # number of segments (2 segments is equal to 1 cycle )
        self.I_sensitivity = float(self.setting['Value'][7]) # sensitivity of current measurement (A). 0.0001 means 0.00105 uA is maximum.
        self.path = os.getcwd()+'/'

        # 保存ファイル名を生成
        filename = "{0}_{1}_{2}-{3}V_{4}Vs-1.xlsx".format(self.date, self.sample_name, self.low_V,self.high_V, self.scan_rate)
        self.filename = duplicate_rename(filename)
        print(filename)

    def initial_settings(self):
        # エクセルファイルの作成
        self.sheet.cell(row = 1, column = 1).value = 'voltage (V)'
        self.sheet.cell(row = 1, column = 2).value = 'current (mA)'
        self.sheet.cell(row = 1, column = 3).value = 'time (sec)'
        self.sheet.cell(row = 1, column = 4).value = datetime.now()

        # ソースメータの設定
        self.keithley.reset() # とりあえずreset
        self.keithley.disable_buffer()
        self.keithley.use_front_terminals()
        self.keithley.apply_voltage() # 電圧印可モード。
        self.keithley.source_voltage_range = 1 # 印可する電圧のレンジ。 1.000 V のレンジで印可。
        self.keithley.source_voltage = 0 # ソース電圧（印可される電圧）の大きさをとりあえず0に。
        self.keithley.enable_source() # 印可を行う（ソース電源の印可を有効にする）。
        self.keithley.measure_current(nplc=0.01, current=0.000105, auto_range=True) # 電流測定。測定に用いる積分時間の長さを0.01 nplcに設定。測定される電流のレンジを
        self.keithley.current_range = self.I_sensitivity
        self.keithley.compliance_current = 0.01 # まず超えないが、一応設定。
        self.keithley.wires = 4 # 4線式で行うことが肝要。理由も説明できるようになるべし。

    def voltage_apply(self,voltage):
        measurement_interval = 1/(self.scan_rate*1000) #余談だが、0.001/scan_rateにすべきではない。小数を用いると桁落ちの危険。
        meas_time = measurement_interval
        time_accuracy = 0.001

        points = int((self.high_V - self.low_V)*1000) # the number of point which should be recorded. I value is recorded at every 0.001 V.
        dV = 0.001 # 掃引電圧の粗さ 兼 （今回の場合は）電流測定間隔
        for i in range(self.segments):
            step = 0
            for step in range (points):
                keithley.source_voltage = voltage
                rest_measurement = time.time() - base_time - meas_time
                while rest_measurement < 0: # 電流反転予定の時間が来るまで待つ。
                    rest_measurement = time.time() - base_time - meas_time
                    time.sleep(time_accuracy)
                self.z.append(time.time()-base_time) #時間記録
                self.x.append(voltage) # 電圧記録
                self.y.append(keithley.current) # 電流記録(A)
                voltage = voltage + dV # 印可電圧の変更（反映されるのは次のループから）
                meas_time = meas_time + measurement_interval # 次の測定点を設定
            dV = dV*(-1) # 掃引方向を変更
        self.keithley.shutdown()

        # エクセルにデータ記入。リストに一旦入れた値をエクセルへ。
        #この方法が毎度毎度エクセルに記録するよりも速いのかは未検証。
        for i in range(segments * points):
            self.sheet.cell(row = 2+i, column = 1).value = self.x[i]
            self.sheet.cell(row = 2+i, column = 2).value = self.y[i]*1000 # 電流をmA単位で記入する。
            self.sheet.cell(row = 2+i, column = 3).value = self.z[i]
        print('voltage_apply_done.')

    def run(self):
        print(self.setting)
        print('----------------------')
        j = input('Correct setting: (y/n) ')
        print('----------------------')
        if j == 'y':
            print('Mode: Cyclic Voltammetry')
        else:
            print('System: closed')
            sys.exit()

        self.initial_settings()
        self.keithley.enable_source()# なぜかここでもoutputをenableしないとoutputがenableにならない
        print('System: enable')
        base_time = time.time() # 測定において0秒となる時刻を取得
        self.voltage_apply(self.start_V)
        print('System:',self.filename)
        self.book.save(self.path + self.filename) # エクセルファイルを filename の名前で、path の場所に保存。
        print('System: finished')
